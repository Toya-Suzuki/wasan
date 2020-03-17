import sys, getopt
import numpy as np
from PIL import Image
from PIL import ImageOps
import warnings
import blobDetector3
from blobDetector3 import SimpleBlobDetector
from dice import dice
import cv2
# original code : http://scikit-image.org/docs/dev/auto_examples/features_detection/plot_blob.html
# other version : https://github.com/TheLaueLab/blob-detection
from math import sqrt
from skimage import data
from skimage.transform import rescale
from skimage.feature import blob_dog, blob_log, blob_doh
from skimage.color import rgb2gray
import matplotlib.pyplot as plt
import os
import openpyxl

def main(argv):
    #excel summary
    book = openpyxl.load_workbook("/home/toya/Research/Wasan/data/summary/evaluate.xlsx")
    sheet = book['パラメータ評価']
    sheet.cell(row=int(argv[5]), column=1).value = str(argv[1])
    #parameter set

    min_sList=[i for i in range(38, 43)]*12             #int
    min_sList2=[i for i in range(35, 45)]*6             #int
    min_sList3=[i for i in range(50, 65)]*4             #int
    max_sList=[i for i in range(35, 65)]*2              #int
    max_sList2=[i for i in range(40, 50)]*6             #int
    max_sList3=[i for i in range(55, 70)]*4             #int
    num_sList=[i for i in range(1, 4)]*20               #int
    num_sList3=[i for i in range(1, 6)]*12              #int
    sigma_List=[i*0.1 for i in range(15, 66, 10)]*12    #float
    thresList=[i*0.01 for i in range(1, 21)]*3          #float
    thresList2=[i*0.1 for i in range(6, 16)]*6          #float
    thresList3=[i*0.001 for i in range(5, 15)]*6        #float
    overList=[i*0.1 for i in range(1, 11)]*6            #float
    overList2=[i*0.1 for i in range(1, 5)]*15           #float
    overList3=[i*0.1 for i in range(3, 8)]*12           #float
    #parameters:
    #argv[1]:COfile
    #argv[2]:method
    #argv[3]:KPfile
    #argv[4]:KPfileAN
    #argv[5]:pageCount
    for p in range(30):
        image_gray=cv2.imread(argv[1], cv2.IMREAD_GRAYSCALE)
        image=image_gray
        image_black=(255-image_gray)

        maskOutputFile = str(argv[3])+"_"+str(min_sList[p])+"-"+str(max_sList[p])+"-"+str(num_sList[p])+"-"+str(thresList[p])+"-"+str(overList[p])+".tif"

        #switch between the different types of blob detector
        if int(argv[2])==0:
            keypoints=SimpleBlobDetector(argv,image)
            #print ("number of keypoints outside "+str(len(keypoints)))

        elif int(argv[2])==1:
            #parameters: http://scikit-image.org/docs/dev/api/skimage.feature.html#skimage.feature.blob_log
            min_s=min_sList[p]
            max_s=max_sList[p]
            num_s=num_sList[p]
            thres=thresList[p]
            over=overList[p]
            print(min_s, max_s, num_s, thres, over)
            blob_List = blob_log(image_black, min_sigma=min_s, max_sigma=max_s, num_sigma=num_s, threshold=thres, overlap=over)
            # Compute radii in the 3rd column.
            blob_List[:, 2] = blob_List[:, 2] * sqrt(2)
        elif int(argv[2])==2:
            #parameters: http://scikit-image.org/docs/dev/api/skimage.feature.html#skimage.feature.blob_dog
            min_s=min_sList2[p]
            max_s=max_sList2[p]
            sigma=sigma_List[p]
            thres=thresList2[p]
            over=overList2[p]
            print(min_s, max_s, sigma, thres, over)
            blob_List = blob_dog(image_black, min_sigma=min_s, max_sigma=max_s, sigma_ratio=sigma, threshold=thres, overlap=over)
            #blob_List = blob_dog(image_black, min_sigma=35, max_sigma=40, sigma_ratio=1.5, threshold=1.0, overlap=0.3)
            blob_List[:, 2] = blob_List[:, 2] * sqrt(2)
        elif int(argv[2])==3:
            #parameters: http://scikit-image.org/docs/dev/api/skimage.feature.html#skimage.feature.blob_doh
            min_s=min_sList3[p]
            max_s=max_sList3[p]
            num_s=num_sList3[p]
            thres=thresList3[p]
            over=overList3[p]
            print(min_s, max_s, num_s, thres, over)
            blob_List = blob_doh(image_gray, min_sigma=min_s, max_sigma=max_s, num_sigma=5, threshold=thres, overlap=over)
            #blob_List = blob_doh(image_gray, min_sigma=65, max_sigma=70, num_sigma=5, threshold=0.005, overlap=0.7)
        else:
            print("Blob detector type not supported: "+argv[3])
            sys.exit(-1)

        #Now write the results to file
        if int(argv[2])==1 or int(argv[2])==2 or int(argv[2])==3:
            for blob in blob_List:
                y, x, r = blob
                cv2.rectangle(image,(int(x-r), int(y-r)), (int(x+r), int(y+r)), (0, 0, 255), 1)

            outputMask=np.ones((image.shape[0],image.shape[1]),np.uint8)
            i=0
            for blob in blob_List:
                y, x, r = blob
                outputMask[int(y-r):int(y+r),int(x-r):int(x+r)]=255
                i=i+1

            invertOutputMask=255-outputMask
            cv2.imwrite("/home/toya/Research/Wasan/data/summary/AllKanjiPositionTest/"+str(maskOutputFile.split("/")[-1]),invertOutputMask)
            
        else:
            for kp in keypoints:
                x = kp.pt[0]
                y = kp.pt[1]
                r = kp.size
                cv2.rectangle(image,(int(x-r), int(y-r)), (int(x+r), int(y+r)), (0, 0, 255), 1)
            cv2.imwrite(argv[2],image)

        ##############################################################################################################################################################

        warnings.simplefilter('ignore', Image.DecompressionBombWarning)

        im1 = Image.open("/home/toya/Research/Wasan/data/summary/AllKanjiPositionTest/"+str(maskOutputFile.split("/")[-1]))
        im2 = Image.open(argv[4])
        # Error control if only one parameter for inverting image is given.
        if len(sys.argv) == 7:
            print("Invert options must be set for both images")
            sys.exit()

        # Inverting input images if is wanted.
        if len(sys.argv) > 6:
            if int(sys.argv[6]) == 1:
                im1 = ImageOps.invert(im1)

            if int(sys.argv[7]) == 1:
                im2 = ImageOps.invert(im2)

        # Image resize for squared images.
        size = 300, 300

        # Converting to b/w.
        gray1 = im1.convert('L')
        im1 = gray1.point(lambda x: 0 if x < 128 else 255, '1')
        gray2 = im2.convert('L')
        im2 = gray2.point(lambda x: 0 if x < 128 else 255, '1')

        # Dice coeficinet computation
        res = dice(im1, im2)
        
        #################################################################################################################################################################

        sheet.cell(row=1, column=p+2).value = str(min_sList[p])+"-"+str(max_sList[p])+"-"+str(num_sList[p])+"-"+str(thresList[p])+"-"+str(overList[p])
        sheet.cell(row=int(argv[5]), column=p+2).value = res
        print("/home/toya/Research/Wasan/data/summary/AllKanjiPositionTest/"+str(maskOutputFile.split("/")[-1]))
        print(res)
        #exit()
    book.save('/home/toya/Research/Wasan/data/summary/evaluate.xlsx')
    book.close()

if __name__ == "__main__":

    main(sys.argv)